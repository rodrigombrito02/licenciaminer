"""Endpoints de Due Diligence ambiental."""

from __future__ import annotations

import io
import logging

import fitz  # pymupdf
from fastapi import APIRouter, File, Query, UploadFile
from pydantic import BaseModel

logger = logging.getLogger(__name__)

from app.components.dd_inventory import (
    LICENCA_DESC,
    LICENCA_TIPOS,
    filtrar_documentos,
    filtrar_requisitos,
    load_inventario,
    load_requisitos,
)
from app.components.dd_scoring import (
    CONFORMIDADE_ESCALA,
    calcular_checklist_completude,
    calcular_conformidade,
    gerar_recomendacoes,
)

router = APIRouter()


@router.get("/due-diligence/license-types")
def list_license_types():
    """Lista tipos de licença disponíveis com descrições."""
    return [
        {"code": code, "description": LICENCA_DESC.get(code, code)}
        for code in LICENCA_TIPOS
    ]


@router.get("/due-diligence/scale")
def get_conformidade_scale():
    """Retorna escala de conformidade (labels, cores, faixas)."""
    return CONFORMIDADE_ESCALA


@router.get("/due-diligence/documents")
def get_documents(
    licenca_tipo: str = Query(..., description="Tipo de licença (LAS, LAS-RAS, LAC1, etc.)"),
):
    """Retorna documentos aplicáveis para um tipo de licença."""
    docs = filtrar_documentos(licenca_tipo)
    return {
        "licenca_tipo": licenca_tipo,
        "total": len(docs),
        "documents": docs,
    }


@router.get("/due-diligence/requirements")
def get_requirements(
    documento: str = Query(..., description="Nome do documento"),
):
    """Retorna requisitos de teste para um documento específico."""
    reqs = filtrar_requisitos(documento)
    return {
        "documento": documento,
        "total": len(reqs),
        "requirements": reqs,
    }


@router.get("/due-diligence/all-requirements")
def get_all_requirements(
    licenca_tipo: str = Query(..., description="Tipo de licença (LAS, LAS-RAS, LAC1, etc.)"),
):
    """Retorna TODOS os requisitos aplicáveis a um tipo de licença.

    Mapeamento explícito: licença → chaves de documento nos requisitos.
    """
    docs = filtrar_documentos(licenca_tipo)
    all_reqs_data = load_requisitos()

    # Explicit mapping: which req-documento keys apply to each license type
    # These match the 'documento' column in dd_requisitos_testes.csv
    LICENCA_REQ_KEYS: dict[str, list[str]] = {
        "LAS": ["Cadastro via SEI"],
        "LAS-RAS": ["Cadastro via SEI", "LAS_RAS"],
        "LAC1": ["Cadastro via SEI", "LAS_RAS", "EIA", "RIMA_LAE", "PCA", "PRAD", "PEA", "PGA", "PIA", "PAFEM", "IDAL", "RADA"],
        "LAC2": ["Cadastro via SEI", "LAS_RAS", "EIA", "PCA", "PRAD", "PEA", "PGA", "PIA", "PAFEM", "IDAL", "RADA"],
        "LP": ["EIA", "PRAD", "PEA", "PGA", "PIA", "PAFEM", "IDAL", "RADA"],
        "LI": ["PCA", "PRAD", "PEA", "PGA", "PIA", "PAFEM", "IDAL"],
        "LO": ["PCA", "PRAD"],
        "LAU": ["RCA_LAU", "PCA_LAU", "PROJ_LAU"],
        "LAC_FED": ["RCE_LAC", "TAC_LAC"],
        "LAE": ["EIA_LAE", "RIMA_LAE", "PBA_LAE"],
        "LOC": ["RCA_LOC", "PCA_LOC"],
    }

    # Collect applicable req-doc keys
    req_doc_keys = set(LICENCA_REQ_KEYS.get(licenca_tipo, []))

    # Also add any doc_ids from inventory that aren't "-" or empty
    for d in docs:
        did = d.get("doc_id", "").strip()
        if did and did != "-":
            req_doc_keys.add(did)

    # Filter requirements
    reqs = [
        r for r in all_reqs_data
        if r.get("documento", "").strip() in req_doc_keys
    ]

    return {
        "licenca_tipo": licenca_tipo,
        "total_documents": len(docs),
        "total_requirements": len(reqs),
        "requirements": reqs,
    }


class ScoreRequest(BaseModel):
    """Payload para cálculo de conformidade."""
    avaliacoes: dict[str, str]
    pesos: dict[str, float] | None = None
    doc_status: dict[str, str] | None = None


@router.post("/due-diligence/score")
def calculate_score(request: ScoreRequest):
    """Calcula score de conformidade a partir das avaliações.

    Aceita avaliações (requisito_id → "Atende" | "Atende Parcialmente" | "Não Atende" | "Não Aplica")
    e opcionalmente pesos e status de documentos.
    """
    resultado = calcular_conformidade(request.avaliacoes, request.pesos)

    response = {
        "total_requisitos": resultado.total_requisitos,
        "requisitos_aplicaveis": resultado.requisitos_aplicaveis,
        "atende": resultado.atende,
        "atende_parcial": resultado.atende_parcial,
        "nao_atende": resultado.nao_atende,
        "nao_aplica": resultado.nao_aplica,
        "conformidade_nao_ponderada": resultado.conformidade_nao_ponderada,
        "conformidade_ponderada": resultado.conformidade_ponderada,
        "nota_maxima": resultado.nota_maxima,
        "nota_obtida": resultado.nota_obtida,
        "classificacao": resultado.classificacao,
        "cor": resultado.cor,
        "descricao": resultado.descricao,
    }

    # Checklist de documentos (se fornecido)
    if request.doc_status:
        response["checklist"] = calcular_checklist_completude(request.doc_status)

    # Recomendações (passa doc_status para identificar documentos ausentes)
    all_reqs = load_requisitos()
    response["recomendacoes"] = gerar_recomendacoes(
        request.avaliacoes, all_reqs, doc_status=request.doc_status
    )

    return response


# ── Upload e extração de documentos ──────────────────────────────────────


@router.post("/due-diligence/upload")
async def upload_document(file: UploadFile = File(...)):
    """Faz upload de PDF e extrai texto para análise.

    Retorna metadados do arquivo e texto extraído (sem persistir no disco).
    """
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        return {"error": "Apenas arquivos PDF são aceitos."}

    content = await file.read()
    size_bytes = len(content)

    try:
        doc = fitz.open(stream=content, filetype="pdf")
        pages = len(doc)
        text_parts = []
        for page in doc:
            text_parts.append(page.get_text())
        doc.close()
        full_text = "\n".join(text_parts)
    except Exception as e:
        logger.warning("Erro ao extrair texto do PDF: %s", e)
        return {"error": f"Erro ao processar PDF: {e}"}

    return {
        "filename": file.filename,
        "pages": pages,
        "size_bytes": size_bytes,
        "text_length": len(full_text),
        "text_preview": full_text[:2000] if full_text else "",
        "extracted_text": full_text,
    }


# ── Criticidade ──────────────────────────────────────────────────────────


class CriticalityRequest(BaseModel):
    """Payload para análise de criticidade."""
    avaliacoes: dict[str, str]


@router.post("/due-diligence/criticality")
def analyze_criticality(request: CriticalityRequest):
    """Calcula matrix de criticidade e aderência por tema.

    Retorna: itens não-conformes com impacto/complexidade + aderência por tópico.
    """
    all_reqs = load_requisitos()

    # Mapear requisito_id → detalhes
    req_map = {r["requisito_id"]: r for r in all_reqs}

    # Itens não-conformes (para a matrix de criticidade)
    non_conformes = []
    for req_id, avaliacao in request.avaliacoes.items():
        if avaliacao in ("Não Atende", "Atende Parcialmente"):
            req = req_map.get(req_id, {})
            impacto_raw = req.get("impacto", "")
            peso_raw = req.get("peso", "")
            # Converter impacto/peso em score numérico (1-5)
            try:
                impacto = float(impacto_raw) if impacto_raw else 3.0
            except (ValueError, TypeError):
                impacto = 3.0
            try:
                complexidade = float(peso_raw) if peso_raw else 2.0
            except (ValueError, TypeError):
                complexidade = 2.0
            # Quadrante
            if impacto >= 3 and complexidade >= 3:
                quadrante = "Ação Imediata (alta complexidade)"
            elif impacto >= 3:
                quadrante = "Ação Imediata (simples)"
            elif complexidade >= 3:
                quadrante = "Ações Secundárias"
            else:
                quadrante = "Baixa Prioridade"

            non_conformes.append({
                "requisito_id": req_id,
                "documento": req.get("documento", ""),
                "topico": req.get("topico", ""),
                "teste": req.get("teste_aderencia", ""),
                "avaliacao": avaliacao,
                "impacto": impacto,
                "complexidade": complexidade,
                "quadrante": quadrante,
            })

    # Aderência por tópico
    topicos: dict[str, dict] = {}
    for req_id, avaliacao in request.avaliacoes.items():
        if avaliacao == "Não Aplica":
            continue
        req = req_map.get(req_id, {})
        topico = req.get("topico", "Sem tópico")
        if topico not in topicos:
            topicos[topico] = {"topico": topico, "total": 0, "atende": 0, "parcial": 0, "nao_atende": 0}
        topicos[topico]["total"] += 1
        if avaliacao == "Atende":
            topicos[topico]["atende"] += 1
        elif avaliacao == "Atende Parcialmente":
            topicos[topico]["parcial"] += 1
        elif avaliacao == "Não Atende":
            topicos[topico]["nao_atende"] += 1

    por_tema = []
    for t in topicos.values():
        if t["total"] > 0:
            t["taxa_aderencia"] = round(100.0 * (t["atende"] + 0.5 * t["parcial"]) / t["total"], 1)
        else:
            t["taxa_aderencia"] = 0
        por_tema.append(t)

    por_tema.sort(key=lambda x: x["taxa_aderencia"])

    # Gargalos (3 temas com menor aderência)
    gargalos = [t["topico"] for t in por_tema[:3] if t["taxa_aderencia"] < 80]

    return {
        "non_conformes": non_conformes,
        "total_non_conformes": len(non_conformes),
        "por_tema": por_tema,
        "gargalos": gargalos,
        "quadrantes": {
            "acao_imediata_complexa": sum(1 for n in non_conformes if "alta complexidade" in n["quadrante"]),
            "acao_imediata_simples": sum(1 for n in non_conformes if "simples" in n["quadrante"]),
            "acoes_secundarias": sum(1 for n in non_conformes if "Secundárias" in n["quadrante"]),
            "baixa_prioridade": sum(1 for n in non_conformes if "Baixa" in n["quadrante"]),
        },
    }


# ── Export XLSX ──────────────────────────────────────────────────────────


class XlsxExportRequest(BaseModel):
    """Payload para exportar planilha de acompanhamento."""
    avaliacoes: dict[str, str]
    doc_status: dict[str, str] | None = None
    licenca_tipo: str = ""


@router.post("/due-diligence/export-xlsx")
def export_xlsx(request: XlsxExportRequest):
    """Gera planilha XLSX de acompanhamento com 4 abas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from starlette.responses import StreamingResponse

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")

    all_reqs = load_requisitos()
    req_map = {r["requisito_id"]: r for r in all_reqs}
    resultado = calcular_conformidade(request.avaliacoes)

    # ── Aba 1: Dashboard ──
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1["A1"] = "Due Diligence - Dashboard"
    ws1["A1"].font = Font(bold=True, size=14, color="0A2540")
    ws1["A3"] = "Score Global"
    ws1["B3"] = f"{round(resultado.conformidade_nao_ponderada * 100, 1)}%"
    ws1["B3"].font = Font(bold=True, size=16)
    ws1["A4"] = "Classificacao"
    ws1["B4"] = resultado.classificacao
    ws1["A5"] = "Total Requisitos"
    ws1["B5"] = resultado.total_requisitos
    ws1["A6"] = "Atende"
    ws1["B6"] = resultado.atende
    ws1["A7"] = "Parcial"
    ws1["B7"] = resultado.atende_parcial
    ws1["A8"] = "Nao Atende"
    ws1["B8"] = resultado.nao_atende
    ws1["A9"] = "Nao Aplica"
    ws1["B9"] = resultado.nao_aplica
    ws1["A11"] = "Licenca"
    ws1["B11"] = request.licenca_tipo
    for row in ws1.iter_rows(min_row=3, max_row=11, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # ── Aba 2: Inventario ──
    ws2 = wb.create_sheet("Inventario")
    inv_headers = ["#", "Documento", "Status", "Arquivo Vinculado"]
    for col, h in enumerate(inv_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    doc_status = request.doc_status or {}
    inventario = load_inventario()
    for i, doc in enumerate(inventario):
        doc_name = doc.get("documento", "")
        status = doc_status.get(doc_name, "Nao Apresentado")
        ws2.cell(row=i+2, column=1, value=i+1)
        ws2.cell(row=i+2, column=2, value=doc_name)
        c = ws2.cell(row=i+2, column=3, value=status)
        if status == "Apresentado":
            c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        elif status == "Nao Apresentado":
            c.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws2.cell(row=i+2, column=4, value="")
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 30
    ws2.auto_filter.ref = f"A1:D{len(inventario)+1}"

    # ── Aba 3: Avaliacao ──
    ws3 = wb.create_sheet("Avaliacao")
    eval_headers = ["ID", "Documento", "Modulo", "Topico", "Teste", "Evidencia", "Resultado", "Observacoes"]
    for col, h in enumerate(eval_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row_num = 2
    for req_id, avaliacao in sorted(request.avaliacoes.items()):
        req = req_map.get(req_id, {})
        ws3.cell(row=row_num, column=1, value=req_id)
        ws3.cell(row=row_num, column=2, value=req.get("documento", ""))
        ws3.cell(row=row_num, column=3, value=req.get("modulo", ""))
        ws3.cell(row=row_num, column=4, value=req.get("topico", ""))
        ws3.cell(row=row_num, column=5, value=req.get("teste_aderencia", ""))
        ws3.cell(row=row_num, column=6, value=req.get("evidencia_esperada", ""))
        c = ws3.cell(row=row_num, column=7, value=avaliacao)
        if avaliacao == "Atende":
            c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        elif "Parcial" in avaliacao:
            c.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        elif "Nao" in avaliacao:
            c.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws3.cell(row=row_num, column=8, value="")
        row_num += 1
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 25
    ws3.column_dimensions["E"].width = 50
    ws3.column_dimensions["F"].width = 40
    ws3.column_dimensions["G"].width = 18
    ws3.column_dimensions["H"].width = 30
    ws3.auto_filter.ref = f"A1:H{row_num-1}"

    # ── Aba 4: Plano de Acao ──
    ws4 = wb.create_sheet("Plano de Acao")
    plan_headers = ["#", "Prioridade", "Tipo", "Documento", "Requisito", "Acao", "Responsavel", "Prazo", "Status"]
    for col, h in enumerate(plan_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    recomendacoes = gerar_recomendacoes(request.avaliacoes, all_reqs, doc_status=request.doc_status)
    for i, rec in enumerate(recomendacoes):
        ws4.cell(row=i+2, column=1, value=i+1)
        c = ws4.cell(row=i+2, column=2, value=rec.get("prioridade", ""))
        if rec.get("prioridade") == "Alta":
            c.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws4.cell(row=i+2, column=3, value=rec.get("tipo", ""))
        ws4.cell(row=i+2, column=4, value=rec.get("documento", ""))
        ws4.cell(row=i+2, column=5, value=rec.get("requisito_id", ""))
        ws4.cell(row=i+2, column=6, value=rec.get("teste", "") or rec.get("evidencia", ""))
        ws4.cell(row=i+2, column=7, value="")  # Responsavel (vazio)
        ws4.cell(row=i+2, column=8, value="")  # Prazo (vazio)
        ws4.cell(row=i+2, column=9, value="")  # Status (vazio)
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 20
    ws4.column_dimensions["D"].width = 20
    ws4.column_dimensions["E"].width = 10
    ws4.column_dimensions["F"].width = 50
    ws4.column_dimensions["G"].width = 20
    ws4.column_dimensions["H"].width = 15
    ws4.column_dimensions["I"].width = 15
    ws4.auto_filter.ref = f"A1:I{len(recomendacoes)+1}"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=DD_Acompanhamento.xlsx"},
    )
