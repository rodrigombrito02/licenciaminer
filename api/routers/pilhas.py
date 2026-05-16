"""Endpoints do módulo Conformidade de Pilhas de Rejeito/Estéril.

Reaproveita o motor de scoring/recomendações do Due Diligence e adiciona
endpoints específicos para pilhas:

  GET  /api/pilhas/stats                   — estatísticas do inventário
  GET  /api/pilhas/normas                  — 23 normas do arcabouço
  GET  /api/pilhas/etapas                  — lista de etapas do ciclo
  GET  /api/pilhas/modalidades             — modalidades aplicáveis
  GET  /api/pilhas/documents               — documentos aplicáveis (filtros)
  GET  /api/pilhas/requirements            — requisitos de um documento
  GET  /api/pilhas/all-requirements        — todos os requisitos aplicáveis
  POST /api/pilhas/score                   — calcula score de conformidade
  POST /api/pilhas/report/conformidade     — gera relatório HTML
"""

from __future__ import annotations

import logging

from fastapi import APIRouter, Query
from pydantic import BaseModel

from app.components.pilhas_inventory import (
    ATIVIDADES_PILHAS,
    ETAPA_DESC,
    ETAPAS,
    ETAPAS_POR_MODO,
    MODALIDADE_DESC,
    MODALIDADE_TIPOS,
    MODO_DESC,
    estatisticas,
    filtrar_documentos,
    filtrar_por_etapa,
    filtrar_por_modo,
    filtrar_requisitos,
    load_normas,
    load_requisitos,
    requisitos_por_modalidade,
    requisitos_por_modo,
)
from app.components.dd_scoring import (
    CONFORMIDADE_ESCALA,
    calcular_conformidade,
    gerar_recomendacoes,
)
from api.services.database import run_query
from licenciaminer.database.queries import (
    QUERY_CNPJ_PROFILE,
    QUERY_CNPJ_ANM_TITULOS,
    QUERY_CNPJ_CFEM,
    QUERY_CNPJ_INFRACOES,
)
from api.services.report_templates import render_pilhas_conformidade, render_pilhas_portal_publico
from fastapi.responses import HTMLResponse, StreamingResponse
import io

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/pilhas/stats")
def get_stats():
    """Estatísticas do inventário (badges de UI)."""
    return estatisticas()


@router.get("/pilhas/normas")
def get_normas():
    """Lista o arcabouço regulatório — 23 normas em 6 eixos."""
    return load_normas()


@router.get("/pilhas/etapas")
def get_etapas():
    """Etapas do ciclo de vida da pilha."""
    return [
        {"codigo": e, "descricao": ETAPA_DESC.get(e, e)}
        for e in ETAPAS
    ]


@router.get("/pilhas/modalidades")
def get_modalidades():
    """Modalidades de licenciamento aplicáveis."""
    return [
        {"code": m, "description": MODALIDADE_DESC.get(m, m)}
        for m in MODALIDADE_TIPOS
    ]


@router.get("/pilhas/modos")
def get_modos():
    """Modos de uso da ferramenta (Auditoria, Licenciamento, Fechamento)."""
    return [
        {
            "codigo": m,
            "descricao": MODO_DESC[m],
            "etapas": ETAPAS_POR_MODO[m],
        }
        for m in ("AUDITORIA", "LICENCIAMENTO", "FECHAMENTO_MODO")
    ]


@router.get("/pilhas/atividades")
def get_atividades():
    """Atividades DN COPAM 217/2017 relacionadas a pilhas."""
    return [
        {"codigo": c, "descricao": d}
        for c, d in ATIVIDADES_PILHAS.items()
    ]


@router.get("/pilhas/documents")
def get_documents(
    modo: str | None = Query(None, description="Modo de uso (AUDITORIA, LICENCIAMENTO, FECHAMENTO_MODO)"),
    modalidade: str | None = Query(None, description="Modalidade (só para modo LICENCIAMENTO)"),
    incluir_gistm: bool = Query(False, description="Incluir módulo GISTM (governança premium)"),
    etapa: str | None = Query(None, description="Filtrar por etapa (PROJETO, OPERACAO, etc.)"),
):
    """Documentos aplicáveis.

    Filtragem principal:
    - `modo`: se fornecido, filtra por conjunto de etapas do modo
      (AUDITORIA = etapas de ativo existente; LICENCIAMENTO = etapas
      pré-protocolo; FECHAMENTO_MODO = etapas de descomissionamento).
    - `modalidade`: alternativa ao modo, para filtragem por licenciamento.
      Se ambos forem omitidos, padrão é `modo=AUDITORIA`.
    - `etapa`: filtro adicional refinando a lista.
    - `incluir_gistm`: adiciona/permite etapa GOVERNANCA_GISTM.
    """
    if modo:
        docs = filtrar_por_modo(modo.upper(), incluir_gistm=incluir_gistm)
    elif modalidade:
        docs = filtrar_documentos(modalidade, incluir_gistm=incluir_gistm)
    else:
        docs = filtrar_por_modo("AUDITORIA", incluir_gistm=incluir_gistm)
        modo = "AUDITORIA"

    if etapa:
        alvo = etapa.strip().upper()
        docs = [d for d in docs if d.get("etapa", "").strip().upper() == alvo]

    return {
        "modo": modo,
        "modalidade": modalidade,
        "incluir_gistm": incluir_gistm,
        "etapa_filtro": etapa,
        "total": len(docs),
        "documents": docs,
    }


@router.get("/pilhas/requirements")
def get_requirements(
    documento: str = Query(..., description="Nome do documento"),
):
    """Requisitos de um documento específico."""
    reqs = filtrar_requisitos(documento)
    return {"documento": documento, "total": len(reqs), "requirements": reqs}


@router.get("/pilhas/all-requirements")
def get_all_requirements(
    modo: str | None = Query(None),
    modalidade: str | None = Query(None),
    incluir_gistm: bool = Query(False),
):
    """Todos os requisitos aplicáveis ao caso.

    Aceita `modo` (AUDITORIA/LICENCIAMENTO/FECHAMENTO_MODO) OU `modalidade`.
    Padrão: `modo=AUDITORIA`.
    """
    if modo:
        reqs = requisitos_por_modo(modo.upper(), incluir_gistm=incluir_gistm)
    elif modalidade:
        reqs = requisitos_por_modalidade(modalidade, incluir_gistm=incluir_gistm)
    else:
        reqs = requisitos_por_modo("AUDITORIA", incluir_gistm=incluir_gistm)
        modo = "AUDITORIA"
    return {
        "modo": modo,
        "modalidade": modalidade,
        "incluir_gistm": incluir_gistm,
        "total": len(reqs),
        "requirements": reqs,
    }


class DadosPilha(BaseModel):
    """Dados da pilha existente (para Modo Auditoria)."""
    nome: str | None = None
    classe: int | None = None  # 1..6 conforme DN COPAM 217
    tipo: str | None = None  # rejeito / esteril / mista
    metodo_construtivo: str | None = None  # dry_stack / empilhamento_drenado / PDE_convencional
    material: str | None = None  # minerio_ferro / ouro / bauxita / etc.
    altura_m: float | None = None
    volume_m3: float | None = None
    data_inicio: str | None = None  # AAAA-MM-DD
    consequencia: str | None = None  # low/significant/high/very_high/extreme (GISTM)
    municipio: str | None = None
    cnpj: str | None = None


class PilhaScoreRequest(BaseModel):
    modo: str = "AUDITORIA"
    modalidade: str | None = None
    incluir_gistm: bool = False
    avaliacoes: dict[str, str]
    doc_status: dict[str, str] = {}
    dados_pilha: DadosPilha | None = None


@router.post("/pilhas/score")
def score_pilhas(request: PilhaScoreRequest):
    """Calcula score de conformidade (reaproveita motor do DD).

    Entrada: modo + avaliações por requisito + (opcional) dados da pilha.
    Saída: classificação, cor, KPIs, recomendações.
    """
    if request.modo and request.modo.upper() != "LICENCIAMENTO":
        reqs = requisitos_por_modo(
            request.modo.upper(), incluir_gistm=request.incluir_gistm
        )
    elif request.modalidade:
        reqs = requisitos_por_modalidade(
            request.modalidade, incluir_gistm=request.incluir_gistm
        )
    else:
        reqs = requisitos_por_modo("AUDITORIA", incluir_gistm=request.incluir_gistm)
    all_reqs_ids = {r["requisito_id"] for r in reqs}

    # normaliza avaliacoes (remove ids fora do escopo)
    aval_filtrada = {
        k: v for k, v in request.avaliacoes.items() if k in all_reqs_ids
    }

    # pesos por requisito (peso do CSV, coluna "peso")
    pesos: dict[str, float] = {}
    for r in reqs:
        try:
            pesos[r["requisito_id"]] = float(r.get("peso") or 1.0)
        except (TypeError, ValueError):
            pesos[r["requisito_id"]] = 1.0

    resultado = calcular_conformidade(aval_filtrada, pesos)
    recomendacoes = gerar_recomendacoes(
        aval_filtrada, reqs, doc_status=request.doc_status
    )

    # resultado é dataclass/pydantic — converter para dict se necessário
    if hasattr(resultado, "__dict__"):
        resultado_dict = resultado.__dict__
    elif hasattr(resultado, "model_dump"):
        resultado_dict = resultado.model_dump()
    else:
        resultado_dict = dict(resultado)

    return {
        **resultado_dict,
        "recomendacoes": recomendacoes,
        "modo": request.modo,
        "modalidade": request.modalidade,
        "incluir_gistm": request.incluir_gistm,
        "dados_pilha": request.dados_pilha.model_dump() if request.dados_pilha else None,
    }


@router.get("/pilhas/scale")
def get_conformidade_scale():
    """Escala de conformidade (cores e faixas)."""
    return CONFORMIDADE_ESCALA


@router.get("/pilhas/lookup-by-cnpj/{cnpj}")
def lookup_pilha_by_cnpj(cnpj: str):
    """Lookup de dados públicos por CNPJ para auto-popular DadosPilha.

    Reaproveita queries da plataforma (perfil empresa + títulos ANM +
    CFEM + infrações) e devolve um pacote pré-mastigado para o frontend
    sugerir campos do form.

    Heurística de "provavelmente opera pilhas":
    - tem ≥ 1 título ANM em fase de Lavra/Concessão
    - paga CFEM ativamente (≥ 6 meses)
    - substância principal é minério de ferro/ouro/bauxita/cobre/zinco
      (atividades A-05 da DN COPAM 217 com pilhas obrigatórias)
    """
    cnpj_clean = "".join(c for c in cnpj if c.isdigit())
    if len(cnpj_clean) != 14:
        return {"erro": "CNPJ inválido — informe 14 dígitos", "cnpj": cnpj}

    profile_rows = run_query(QUERY_CNPJ_PROFILE, [cnpj_clean])
    profile = profile_rows[0] if profile_rows else None
    if not profile:
        return {
            "cnpj": cnpj_clean,
            "encontrado": False,
            "mensagem": "CNPJ não encontrado nas bases públicas (ANM/IBAMA/CFEM).",
        }

    razao = profile.get("razao_social") or ""
    titulos = run_query(QUERY_CNPJ_ANM_TITULOS, [razao]) if razao else []
    cfem_rows = run_query(QUERY_CNPJ_CFEM, [cnpj_clean])
    infracoes_rows = run_query(QUERY_CNPJ_INFRACOES, [cnpj_clean])

    # CFEM breakdown por substância para identificar material principal
    cfem_subst = run_query(
        """
        SELECT
            "Substância" AS substancia,
            "Município" AS municipio,
            COUNT(*) AS n_pagamentos
        FROM v_cfem
        WHERE CPF_CNPJ = ?
        GROUP BY 1, 2
        ORDER BY n_pagamentos DESC
        LIMIT 10
        """,
        [cnpj_clean],
    )

    # Heurística pilha
    cfem_meses = (cfem_rows[0].get("meses_pagamento", 0) if cfem_rows else 0) or 0
    n_titulos_lavra = sum(
        1 for t in titulos
        if any(k in (t.get("fase") or "").lower()
               for k in ("lavra", "concessão", "concessao"))
    )
    substancias = [(s.get("substancia") or "").lower() for s in cfem_subst]
    SUBST_PILHAS = {"ferro", "ouro", "bauxita", "cobre", "zinco", "manganês", "manganes",
                    "níquel", "niquel", "fosfato", "fosfático", "fosfatico"}
    tem_subst_pilha = any(any(p in s for p in SUBST_PILHAS) for s in substancias)
    # Heurística: opera pilhas se há produção em curso (CFEM ativo) em substância
    # que normativamente gera pilhas, OU se tem título de lavra em substância pilha
    provavel_opera_pilhas = tem_subst_pilha and (cfem_meses >= 12 or n_titulos_lavra >= 1)

    # Sugestão de auto-populate
    sugestao = {}
    if cfem_subst:
        sugestao["material"] = cfem_subst[0].get("substancia")
        sugestao["municipio"] = cfem_subst[0].get("municipio")

    return {
        "cnpj": cnpj_clean,
        "encontrado": True,
        "empresa": {
            "razao_social": razao,
            "municipio_sede": profile.get("municipio"),
            "uf_sede": profile.get("uf"),
            "situacao": profile.get("situacao_cadastral"),
        },
        "titulos_anm": {
            "total": len(titulos),
            "lavra_concessao": n_titulos_lavra,
            "amostra": titulos[:5],
        },
        "cfem": {
            "meses_pagamento": cfem_meses,
            "total_pago": (cfem_rows[0].get("total_pago", 0) if cfem_rows else 0) or 0,
            "substancias_top": cfem_subst[:5],
        },
        "infracoes": {
            "total": (infracoes_rows[0].get("total_infracoes", 0) if infracoes_rows else 0) or 0,
            "anos_com_infracao": (infracoes_rows[0].get("anos_com_infracao", 0) if infracoes_rows else 0) or 0,
        },
        "analise_pilhas": {
            "provavel_opera_pilhas": provavel_opera_pilhas,
            "criterios": {
                "cfem_meses_>=6": cfem_meses >= 6,
                "titulos_lavra_>=1": n_titulos_lavra >= 1,
                "substancia_pilha_obrigatoria": tem_subst_pilha,
            },
        },
        "sugestao_auto_populate": sugestao,
    }


@router.get("/pilhas/gistm-principles")
def get_gistm_principles():
    """Lista os 15 princípios GISTM com requisitos mapeados em cada um.

    Retorna a estrutura usada pelo overlay premium para roll-up de scoring
    por princípio (e não apenas por requisito individual).
    """
    all_reqs = load_requisitos()
    gistm_reqs = [r for r in all_reqs if "GISTM" in (r.get("modulo") or "")]

    principles: dict[int, dict] = {}
    for r in gistm_reqs:
        topico = r.get("topico", "")
        # Formato esperado: "P01 - Engajamento..."
        if not topico.startswith("P"):
            continue
        try:
            pnum = int(topico[1:3])
        except ValueError:
            continue
        if pnum not in principles:
            principles[pnum] = {
                "principio": pnum,
                "nome": topico.split(" - ", 1)[1] if " - " in topico else topico,
                "requisitos": [],
            }
        principles[pnum]["requisitos"].append({
            "requisito_id": r.get("requisito_id"),
            "teste": r.get("teste_aderencia"),
            "peso": r.get("peso"),
            "impacto": r.get("impacto"),
        })

    return {
        "total_principios": len(principles),
        "total_requisitos": len(gistm_reqs),
        "principios": [principles[k] for k in sorted(principles.keys())],
    }


class GistmScoreRequest(BaseModel):
    avaliacoes: dict[str, str]


@router.post("/pilhas/gistm-score")
def score_gistm_by_principle(request: GistmScoreRequest):
    """Calcula score por princípio GISTM a partir das avaliações.

    Útil para o overlay premium: mostra rating de aderência aos 15 princípios
    independentemente do score global do modo.
    """
    all_reqs = load_requisitos()
    gistm_reqs = [r for r in all_reqs if "GISTM" in (r.get("modulo") or "")]

    per_principle: dict[int, dict] = {}
    for r in gistm_reqs:
        topico = r.get("topico", "")
        if not topico.startswith("P"):
            continue
        try:
            pnum = int(topico[1:3])
        except ValueError:
            continue
        if pnum not in per_principle:
            per_principle[pnum] = {
                "principio": pnum,
                "nome": topico.split(" - ", 1)[1] if " - " in topico else topico,
                "atende": 0, "parcial": 0, "nao_atende": 0, "nao_aplica": 0,
                "nao_avaliado": 0, "total": 0, "score_pct": None,
            }
        bucket = per_principle[pnum]
        bucket["total"] += 1
        aval = request.avaliacoes.get(r["requisito_id"], "")
        if aval == "Atende":
            bucket["atende"] += 1
        elif "Parcial" in aval:
            bucket["parcial"] += 1
        elif aval.startswith("Nao Atende") or aval == "Não Atende":
            bucket["nao_atende"] += 1
        elif "Aplica" in aval and "Nao" in aval or aval == "Não Aplica":
            bucket["nao_aplica"] += 1
        else:
            bucket["nao_avaliado"] += 1

    for p in per_principle.values():
        avaliados = p["total"] - p["nao_aplica"] - p["nao_avaliado"]
        if avaliados > 0:
            p["score_pct"] = round(100.0 * (p["atende"] + 0.5 * p["parcial"]) / avaliados, 1)

    # Score global GISTM (média ponderada simples)
    scores_validos = [p["score_pct"] for p in per_principle.values() if p["score_pct"] is not None]
    global_score = round(sum(scores_validos) / len(scores_validos), 1) if scores_validos else None

    return {
        "score_global_gistm": global_score,
        "principios_avaliados": len(scores_validos),
        "total_principios": len(per_principle),
        "principios": [per_principle[k] for k in sorted(per_principle.keys())],
    }


@router.post("/pilhas/report/conformidade", response_class=HTMLResponse)
def gerar_relatorio_conformidade(request: PilhaScoreRequest):
    """Gera relatório HTML de conformidade da pilha.

    Recebe o mesmo payload do /pilhas/score, calcula o resultado e
    renderiza relatório pronto para impressão (Ctrl+P → PDF).
    """
    modo = (request.modo or "AUDITORIA").upper()
    if modo != "LICENCIAMENTO":
        reqs = requisitos_por_modo(modo, incluir_gistm=request.incluir_gistm)
    elif request.modalidade:
        reqs = requisitos_por_modalidade(request.modalidade, incluir_gistm=request.incluir_gistm)
    else:
        reqs = requisitos_por_modo("AUDITORIA", incluir_gistm=request.incluir_gistm)

    all_ids = {r["requisito_id"] for r in reqs}
    aval_filtrada = {k: v for k, v in request.avaliacoes.items() if k in all_ids}
    pesos: dict[str, float] = {}
    for r in reqs:
        try:
            pesos[r["requisito_id"]] = float(r.get("peso") or 1.0)
        except (TypeError, ValueError):
            pesos[r["requisito_id"]] = 1.0

    resultado = calcular_conformidade(aval_filtrada, pesos)
    recomendacoes = gerar_recomendacoes(aval_filtrada, reqs, doc_status=request.doc_status)

    if hasattr(resultado, "__dict__"):
        resultado_dict = resultado.__dict__
    elif hasattr(resultado, "model_dump"):
        resultado_dict = resultado.model_dump()
    else:
        resultado_dict = dict(resultado)

    # GISTM overlay quando solicitado
    gistm_data = None
    if request.incluir_gistm:
        gistm_data = score_gistm_by_principle(GistmScoreRequest(avaliacoes=request.avaliacoes))

    html = render_pilhas_conformidade(
        modo=modo,
        modo_desc=MODO_DESC.get(modo, modo),
        dados_pilha=request.dados_pilha.model_dump() if request.dados_pilha else None,
        resultado=resultado_dict,
        recomendacoes=recomendacoes or [],
        incluir_gistm=request.incluir_gistm,
        gistm_data=gistm_data,
    )
    return HTMLResponse(content=html)


class PortalPublicoRequest(BaseModel):
    """Payload para gerar pagina publica de transparencia (PL 2.519/3.799)."""
    dados_pilha: DadosPilha
    avaliacoes: dict[str, str] = {}
    incluir_gistm: bool = False
    empresa: dict | None = None  # {"razao_social": "...", "municipio_sede": "..."}


@router.post("/pilhas/portal-publico", response_class=HTMLResponse)
def gerar_portal_publico(request: PortalPublicoRequest):
    """Gera pagina publica de transparencia para a pilha (PL 2.519 MG / PL 3.799 Fed).

    Diferente do relatorio confidencial: linguagem acessivel a comunidade,
    foco em seguranca e emergencia, omite percentuais sensiveis. Pronta para
    publicacao no portal institucional da empresa operadora.

    Quando avaliacoes sao fornecidas, indica nivel de gestao (em conformidade /
    em adequacao / em remediacao) sem expor score numerico. Quando incluir_gistm
    e true, adiciona bloco de adesao ao padrao internacional.
    """
    resultado_dict = None
    if request.avaliacoes:
        # Calcula resultado apenas para classificar nivel (sem expor %)
        all_reqs = load_requisitos()
        all_ids = {r["requisito_id"] for r in all_reqs}
        aval = {k: v for k, v in request.avaliacoes.items() if k in all_ids}
        if aval:
            pesos = {r["requisito_id"]: 1.0 for r in all_reqs}
            res = calcular_conformidade(aval, pesos)
            if hasattr(res, "__dict__"):
                resultado_dict = res.__dict__
            elif hasattr(res, "model_dump"):
                resultado_dict = res.model_dump()
            else:
                resultado_dict = dict(res)

    gistm_data = None
    if request.incluir_gistm and request.avaliacoes:
        gistm_data = score_gistm_by_principle(GistmScoreRequest(avaliacoes=request.avaliacoes))

    html = render_pilhas_portal_publico(
        dados_pilha=request.dados_pilha.model_dump(),
        resultado=resultado_dict,
        gistm_data=gistm_data,
        empresa=request.empresa,
    )
    return HTMLResponse(content=html)


@router.post("/pilhas/export-xlsx")
def export_pilhas_xlsx(request: PilhaScoreRequest):
    """Gera planilha XLSX de auditoria de pilha com 4 abas:
    Dashboard · Inventário de Documentos · Avaliação de Requisitos · Plano de Ação.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    modo = (request.modo or "AUDITORIA").upper()
    incluir_gistm = request.incluir_gistm

    # Carrega documentos e requisitos do escopo do modo
    if modo != "LICENCIAMENTO":
        docs = filtrar_por_modo(modo, incluir_gistm=incluir_gistm)
        reqs = requisitos_por_modo(modo, incluir_gistm=incluir_gistm)
    else:
        from app.components.pilhas_inventory import filtrar_documentos
        docs = filtrar_documentos(request.modalidade or "LAC1", incluir_gistm=incluir_gistm)
        reqs = requisitos_por_modalidade(request.modalidade or "LAC1", incluir_gistm=incluir_gistm)

    req_map = {r["requisito_id"]: r for r in reqs}
    all_ids = set(req_map.keys())
    aval_filtrada = {k: v for k, v in request.avaliacoes.items() if k in all_ids}

    pesos: dict[str, float] = {}
    for r in reqs:
        try:
            pesos[r["requisito_id"]] = float(r.get("peso") or 1.0)
        except (TypeError, ValueError):
            pesos[r["requisito_id"]] = 1.0

    resultado = calcular_conformidade(aval_filtrada, pesos)
    recomendacoes = gerar_recomendacoes(aval_filtrada, reqs, doc_status=request.doc_status)

    if hasattr(resultado, "__dict__"):
        res = resultado.__dict__
    else:
        res = dict(resultado) if not hasattr(resultado, "model_dump") else resultado.model_dump()

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    title_font = Font(bold=True, size=14, color="0A2540")

    dp = request.dados_pilha.model_dump() if request.dados_pilha else {}

    # ── Aba 1: Dashboard ──
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1["A1"] = "Auditoria de Pilha - Dashboard"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:C1")

    rows = [
        ("Pilha", dp.get("nome") or "-"),
        ("Modo", modo),
        ("Modalidade", request.modalidade or "-"),
        ("GISTM Premium", "Sim" if incluir_gistm else "Nao"),
        ("Classe (DN COPAM 217)", dp.get("classe") or "-"),
        ("Tipo", dp.get("tipo") or "-"),
        ("Metodo construtivo", dp.get("metodo_construtivo") or "-"),
        ("Material", dp.get("material") or "-"),
        ("Altura (m)", dp.get("altura_m") or "-"),
        ("Volume (m3)", dp.get("volume_m3") or "-"),
        ("Municipio", dp.get("municipio") or "-"),
        ("Consequencia (GISTM)", dp.get("consequencia") or "-"),
        ("CNPJ", dp.get("cnpj") or "-"),
        ("", ""),
        ("Score Global", f"{round((res.get('conformidade_nao_ponderada', 0) or 0) * 100, 1)}%"),
        ("Classificacao", res.get("classificacao", "-")),
        ("Total Requisitos Aplicaveis", res.get("requisitos_aplicaveis", len(aval_filtrada))),
        ("Atende", res.get("atende", 0)),
        ("Atende Parcialmente", res.get("atende_parcial", 0)),
        ("Nao Atende", res.get("nao_atende", 0)),
        ("Nao Aplica", res.get("nao_aplica", 0)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=v)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 36

    # ── Aba 2: Inventario ──
    ws2 = wb.create_sheet("Inventario")
    inv_headers = ["#", "Etapa", "Documento", "Esfera", "Norma", "Status", "Arquivo Vinculado"]
    for col, h in enumerate(inv_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    ds = request.doc_status or {}
    for i, doc in enumerate(docs, start=2):
        nome = doc.get("documento", "")
        status = ds.get(nome, "Pendente")
        ws2.cell(row=i, column=1, value=i - 1)
        ws2.cell(row=i, column=2, value=doc.get("etapa", ""))
        ws2.cell(row=i, column=3, value=nome)
        ws2.cell(row=i, column=4, value=doc.get("esfera", ""))
        ws2.cell(row=i, column=5, value=doc.get("norma_referencia", ""))
        c = ws2.cell(row=i, column=6, value=status)
        if status.lower() == "apresentado":
            c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        elif status.lower() == "ausente":
            c.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws2.cell(row=i, column=7, value="")
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 24
    ws2.column_dimensions["F"].width = 16
    ws2.column_dimensions["G"].width = 32
    ws2.auto_filter.ref = f"A1:G{max(2, len(docs) + 1)}"

    # ── Aba 3: Avaliacao ──
    ws3 = wb.create_sheet("Avaliacao")
    eval_headers = ["ID", "Modulo", "Topico", "Documento", "Teste de Aderencia", "Evidencia Esperada", "Peso", "Resultado", "Observacoes"]
    for col, h in enumerate(eval_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row_num = 2
    for r in reqs:
        rid = r.get("requisito_id", "")
        avaliacao = aval_filtrada.get(rid, "")
        ws3.cell(row=row_num, column=1, value=rid)
        ws3.cell(row=row_num, column=2, value=r.get("modulo", ""))
        ws3.cell(row=row_num, column=3, value=r.get("topico", ""))
        ws3.cell(row=row_num, column=4, value=r.get("documento", ""))
        ws3.cell(row=row_num, column=5, value=r.get("teste_aderencia", ""))
        ws3.cell(row=row_num, column=6, value=r.get("evidencia_esperada", ""))
        ws3.cell(row=row_num, column=7, value=r.get("peso", ""))
        c = ws3.cell(row=row_num, column=8, value=avaliacao)
        if avaliacao == "Atende":
            c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        elif "Parcial" in avaliacao:
            c.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        elif avaliacao.startswith("Nao Atende"):
            c.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws3.cell(row=row_num, column=9, value="")
        row_num += 1
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 32
    ws3.column_dimensions["E"].width = 55
    ws3.column_dimensions["F"].width = 32
    ws3.column_dimensions["G"].width = 8
    ws3.column_dimensions["H"].width = 18
    ws3.column_dimensions["I"].width = 30
    ws3.auto_filter.ref = f"A1:I{max(2, row_num - 1)}"

    # ── Aba 4: Plano de Acao ──
    ws4 = wb.create_sheet("Plano de Acao")
    plan_headers = ["#", "Prioridade", "Tipo", "Modulo", "Documento", "Requisito ID", "Acao Recomendada", "Responsavel", "Prazo", "Status"]
    for col, h in enumerate(plan_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, rec in enumerate(recomendacoes, start=2):
        ws4.cell(row=i, column=1, value=i - 1)
        c = ws4.cell(row=i, column=2, value=rec.get("prioridade", ""))
        if rec.get("prioridade") == "Alta":
            c.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        elif rec.get("prioridade") == "Media":
            c.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        ws4.cell(row=i, column=3, value=rec.get("tipo", ""))
        rid = rec.get("requisito_id", "")
        modulo = req_map.get(rid, {}).get("modulo", "")
        ws4.cell(row=i, column=4, value=modulo)
        ws4.cell(row=i, column=5, value=rec.get("documento", ""))
        ws4.cell(row=i, column=6, value=rid)
        ws4.cell(row=i, column=7, value=rec.get("teste", "") or rec.get("evidencia", ""))
        # Responsavel/Prazo/Status em branco para o cliente preencher
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 28
    ws4.column_dimensions["E"].width = 32
    ws4.column_dimensions["F"].width = 12
    ws4.column_dimensions["G"].width = 60
    ws4.column_dimensions["H"].width = 22
    ws4.column_dimensions["I"].width = 14
    ws4.column_dimensions["J"].width = 14
    ws4.auto_filter.ref = f"A1:J{max(2, len(recomendacoes) + 1)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = "Auditoria_Pilha"
    if dp.get("nome"):
        slug = "".join(c if c.isalnum() else "_" for c in dp["nome"])[:40]
        nome_arquivo = f"Auditoria_{slug}"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}.xlsx"},
    )
