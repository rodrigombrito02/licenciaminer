"""Parser inteligente de planos heterogêneos (XLSX/CSV).

Lê arquivo, infere mapeamento de colunas pro schema canônico (heurística PT/EN),
extrai EAP, normaliza datas BR/EN e cria registros de TarefaPA.

O sistema é flexível: colunas não mapeadas vão pro `raw_extra` (JSON),
ficando disponíveis para filtro/exibição no dashboard sem schema migration.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

logger = logging.getLogger(__name__)

# ── Dicionário de termos por campo canônico (PT-BR + EN) ──
CANONICAL_TERMS: dict[str, list[str]] = {
    "descricao": [
        "descricao", "descrição", "tarefa", "atividade", "acao", "ação",
        "item", "nome", "titulo", "título", "description", "task", "name",
        "what", "wbs name", "activity", "activity name", "task name",
    ],
    "data_inicio": [
        "data inicio", "data início", "início", "inicio", "data de inicio",
        "data de início", "start", "start date", "begin", "start_date",
        "data prevista inicio", "early start", "planejado inicio",
    ],
    "data_fim": [
        "data fim", "fim", "data de fim", "termino", "término", "data termino",
        "prazo", "deadline", "finish", "finish date", "end", "end date",
        "data limite", "due", "due date", "completion", "data prevista fim",
        "early finish", "planejado fim", "vencimento",
    ],
    "responsavel_pessoa": [
        "responsavel", "responsável", "resp", "owner", "responsible",
        "assigned to", "assignee", "atribuido", "atribuído", "executor",
        "responsavel pessoa", "pessoa responsável", "quem", "who",
    ],
    "area_responsavel": [
        "area", "área", "area responsavel", "área responsável",
        "setor", "departamento", "dept", "department", "diretoria",
        "gerencia", "gerência", "equipe", "time", "team", "squad",
        "function", "unidade", "centro de custo",
    ],
    "status": [
        "status", "situacao", "situação", "estado", "andamento",
        "state", "phase", "stage", "etapa atual",
    ],
    "classificacao": [
        "classificacao", "classificação", "tipo", "categoria", "type",
        "category", "kind", "natureza", "modalidade", "prioridade",
        "priority", "criticidade",
    ],
    "eap_codigo": [
        "eap", "wbs", "estrutura", "codigo eap", "código eap",
        "wbs code", "wbs id", "id eap", "outline number", "outline",
        "nivel", "nível", "hierarquia", "codigo wbs",
    ],
    "pct_concluido": [
        "% concluido", "% concluído", "pct", "percentual", "% complete",
        "complete %", "concluido %", "%complete", "progress", "progresso",
        "% realizado", "realizado %", "avanco", "avanço",
    ],
}


def _normalize_header(s: str) -> str:
    """Normaliza string para matching (lowercase, sem acentos, sem pontuação)."""
    import unicodedata
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^\w\s%]", " ", s.lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def detect_column_mapping(headers: list[str]) -> dict[str, str | None]:
    """Heurística: para cada campo canônico, encontra a melhor coluna do header.

    Retorna {campo_canonico: nome_original_header | None}.
    """
    normalized_headers = {_normalize_header(h): h for h in headers if h}
    mapping: dict[str, str | None] = {}

    for canonical, terms in CANONICAL_TERMS.items():
        match: str | None = None
        # 1) match exato
        for term in terms:
            t = _normalize_header(term)
            if t in normalized_headers:
                match = normalized_headers[t]
                break
        # 2) match por substring (mais permissivo)
        if match is None:
            for term in terms:
                t = _normalize_header(term)
                for nh, orig in normalized_headers.items():
                    if t in nh or nh in t:
                        match = orig
                        break
                if match:
                    break
        mapping[canonical] = match

    return mapping


def _parse_date(value: Any) -> date | None:
    """Aceita date, datetime, string BR/EN/ISO. None se vazio ou inválido."""
    if value is None or value == "" or (isinstance(value, float) and value != value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "null", "-"):
        return None
    # Tentativas: ISO, BR, EN
    for fmt in (
        "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%m/%d/%Y", "%m-%d-%Y",
        "%d/%m/%y", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_eap(value: Any) -> tuple[str | None, int | None, str | None]:
    """Extrai (codigo, nivel, parent) de um valor EAP do tipo "1.2.3" ou "1.2.3.4"."""
    if value is None or value == "":
        return None, None, None
    s = str(value).strip()
    if not re.match(r"^[\d]+(\.[\d]+)*$", s):
        return s, None, None  # outro formato — mantém como string
    parts = s.split(".")
    parent = ".".join(parts[:-1]) if len(parts) > 1 else None
    return s, len(parts), parent


def _parse_pct(value: Any) -> float | None:
    """Aceita 0-1, 0-100, ou string com %. Devolve 0-100."""
    if value is None or value == "":
        return None
    s = str(value).replace("%", "").replace(",", ".").strip()
    if not s:
        return None
    try:
        v = float(s)
        if 0 <= v <= 1:
            v = v * 100
        return round(v, 1)
    except ValueError:
        return None


def parse_xlsx_to_tarefas(
    file_bytes: bytes,
    sheet_name: str | None = None,
    custom_mapping: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Parseia XLSX em lista de tarefas canônicas + raw_extra.

    Args:
        file_bytes: conteúdo do arquivo XLSX
        sheet_name: aba específica; None = primeira aba com dados
        custom_mapping: override do mapeamento automático ({canonico: nome_coluna_original})

    Returns:
        dict com:
          - headers (list[str])
          - mapping (dict[campo_canonico, nome_coluna_original | None])
          - mapping_inferido (bool por campo: True = automatico, False = manual ou None)
          - tarefas (list[dict] pronto pra criar TarefaPA)
          - n_linhas_total, n_linhas_validas
          - sheet_name_usada
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)

    # Escolher aba
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Primeira aba com >= 2 linhas
        ws = None
        for sn in wb.sheetnames:
            cand = wb[sn]
            if cand.max_row >= 2:
                ws = cand
                sheet_name = sn
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]
            sheet_name = wb.sheetnames[0]

    # Detectar linha do header (primeira linha não-vazia que tenha pelo menos 2 strings)
    header_row_idx = 1
    header_row: list[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2 and sum(1 for c in row if isinstance(c, str)) >= 2:
            header_row_idx = i
            header_row = [str(c).strip() if c is not None else "" for c in row]
            break

    if not header_row:
        return {
            "headers": [], "mapping": {}, "mapping_inferido": {},
            "tarefas": [], "n_linhas_total": 0, "n_linhas_validas": 0,
            "sheet_name_usada": sheet_name,
            "erro": "Nao foi possivel detectar cabecalho",
        }

    # Mapping
    auto_mapping = detect_column_mapping(header_row)
    mapping: dict[str, str | None] = dict(auto_mapping)
    if custom_mapping:
        for k, v in custom_mapping.items():
            if v:
                mapping[k] = v
    mapping_inferido = {k: (custom_mapping is None or k not in custom_mapping) for k in CANONICAL_TERMS}

    # Index dos campos
    col_idx: dict[str, int | None] = {}
    for canonical, header_name in mapping.items():
        if header_name and header_name in header_row:
            col_idx[canonical] = header_row.index(header_name)
        else:
            col_idx[canonical] = None

    # Iterar linhas de dados
    tarefas: list[dict] = []
    n_total = 0
    n_validas = 0

    for i, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1
    ):
        n_total += 1
        # Skip linhas totalmente vazias
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        # Truncar row ao tamanho do header (Excel pode trazer celulas vazias extras)
        row_trunc = row[: len(header_row)]

        descricao = (
            str(row_trunc[col_idx["descricao"]]).strip()
            if col_idx["descricao"] is not None and col_idx["descricao"] < len(row_trunc)
            and row_trunc[col_idx["descricao"]] is not None
            else None
        )

        # Pular linha sem descrição (não vale a pena guardar)
        if not descricao:
            continue
        n_validas += 1

        def _get(canonical: str) -> Any:
            idx = col_idx[canonical]
            if idx is None or idx >= len(row_trunc):
                return None
            return row_trunc[idx]

        eap_raw = _get("eap_codigo")
        eap_cod, eap_niv, eap_par = _parse_eap(eap_raw)

        tarefa = {
            "ordem": n_validas,
            "descricao": descricao[:1000],
            "data_inicio": _parse_date(_get("data_inicio")),
            "data_fim": _parse_date(_get("data_fim")),
            "responsavel_pessoa": (str(_get("responsavel_pessoa")).strip()[:200] if _get("responsavel_pessoa") else None),
            "area_responsavel": (str(_get("area_responsavel")).strip()[:200] if _get("area_responsavel") else None),
            "status": (str(_get("status")).strip()[:80] if _get("status") else None),
            "classificacao": (str(_get("classificacao")).strip()[:120] if _get("classificacao") else None),
            "eap_codigo": eap_cod,
            "eap_nivel": eap_niv,
            "parent_eap": eap_par,
            "pct_concluido": _parse_pct(_get("pct_concluido")),
        }

        # raw_extra: tudo que não foi mapeado
        mapped_cols = {idx for idx in col_idx.values() if idx is not None}
        raw_extra = {}
        for j, h in enumerate(header_row):
            if j in mapped_cols or not h:
                continue
            if j >= len(row_trunc):
                continue
            val = row_trunc[j]
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            if isinstance(val, (date, datetime)):
                val = val.isoformat()
            raw_extra[h] = str(val)[:500] if not isinstance(val, (int, float, bool)) else val

        if raw_extra:
            tarefa["raw_extra"] = raw_extra

        tarefas.append(tarefa)

    wb.close()

    return {
        "headers": header_row,
        "mapping": mapping,
        "mapping_inferido": {
            k: (mapping[k] is not None and (custom_mapping is None or k not in (custom_mapping or {})))
            for k in CANONICAL_TERMS
        },
        "tarefas": tarefas,
        "n_linhas_total": n_total,
        "n_linhas_validas": n_validas,
        "sheet_name_usada": sheet_name,
    }
