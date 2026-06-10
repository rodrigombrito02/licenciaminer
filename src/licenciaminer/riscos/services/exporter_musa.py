"""Export Excel do bowtie em layout visual estruturado.

Layout da aba Bowtie:
  Col A | Col B | Col C | D | Col E | F | Col G | Col H | Col I
  CAUSA | BARR  | AÇÃO  |   | TOP   |   | IMPAC | BARR  | AÇÃO
        | PREV  | PREV  |   | EVENT |   | TO    | CORR  | CORR

Risco no centro (mergeado vertical), causas/barreiras/ações preventivas à
esquerda, impactos/barreiras/ações corretivas à direita. Para cada causa,
suas barreiras preventivas e ações preventivas aparecem na mesma linha (ou
linhas empilhadas se forem múltiplas). Espelho para consequências.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from licenciaminer.riscos.models import Acao, Bowtie, Risco


HEADER_DARK = PatternFill("solid", fgColor="1f2937")
HEADER_LIGHT = PatternFill("solid", fgColor="334155")
CAUSA_FILL = PatternFill("solid", fgColor="dbeafe")
BARRIER_PREV_FILL = PatternFill("solid", fgColor="d1fae5")
ACAO_PREV_FILL = PatternFill("solid", fgColor="dcfce7")
CONS_FILL = PatternFill("solid", fgColor="fee2e2")
BARRIER_CORR_FILL = PatternFill("solid", fgColor="ffedd5")
ACAO_CORR_FILL = PatternFill("solid", fgColor="fef3c7")
TOP_EVENT_FILL = PatternFill("solid", fgColor="f59e0b")
CRITICAL_BORDER = Border(
    left=Side(style="medium", color="dc2626"),
    right=Side(style="medium", color="dc2626"),
    top=Side(style="medium", color="dc2626"),
    bottom=Side(style="medium", color="dc2626"),
)
THIN = Side(style="thin", color="94a3b8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WHITE_BOLD = Font(color="ffffff", bold=True, size=11)
WHITE_BIG = Font(color="ffffff", bold=True, size=14)
BLACK_BOLD = Font(bold=True, size=10)

WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _set(ws, row: int, col: int, value, fill=None, font=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    return cell


def _header_row(ws, row: int, cols: list[tuple[int, str]]):
    for col_idx, label in cols:
        _set(ws, row, col_idx, label, fill=HEADER_DARK, font=WHITE_BOLD, align=WRAP_CENTER, border=BOX)
    ws.row_dimensions[row].height = 22


def exportar_bowtie_excel(db: Session, risco_id: int) -> bytes:
    risco = db.get(Risco, risco_id)
    if not risco:
        raise ValueError(f"Risco {risco_id} não encontrado")
    bowtie = (
        db.query(Bowtie)
        .filter_by(risco_id=risco_id)
        .order_by(Bowtie.versao.desc())
        .first()
    )
    acoes_prev = (
        db.query(Acao)
        .filter(Acao.risco_id == risco_id, Acao.tipo == "preventiva")
        .order_by(Acao.codigo, Acao.id)
        .all()
    )
    acoes_corr = (
        db.query(Acao)
        .filter(Acao.risco_id == risco_id, Acao.tipo == "corretiva")
        .order_by(Acao.codigo, Acao.id)
        .all()
    )

    wb = Workbook()

    # ------ Aba 1: Resumo do risco ------
    ws_meta = wb.active
    ws_meta.title = "Risco"
    ws_meta["A1"] = f"{risco.codigo} — {risco.nome}"
    ws_meta["A1"].font = Font(bold=True, size=14, color="1e3a8a")
    ws_meta.merge_cells("A1:D1")

    meta_rows = [
        ("Descrição", risco.descricao or ""),
        ("Estágio", risco.estagio or ""),
        ("Categoria", risco.categoria.nome if risco.categoria else ""),
        ("Responsável", risco.responsavel.nome if risco.responsavel else ""),
        ("Unidade organizacional", risco.unidade_org.nome if risco.unidade_org else ""),
        ("Elo da cadeia de valor", risco.elo_cadeia_valor.nome if risco.elo_cadeia_valor else ""),
        ("Probabilidade pura", risco.prob_pura),
        ("Impacto puro", risco.impacto_pura),
        ("Classificação pura", risco.classificacao_pura or ""),
        ("Probabilidade residual", risco.prob_residual),
        ("Impacto residual", risco.impacto_residual),
        ("Classificação residual", risco.classificacao_residual or ""),
    ]
    for i, (label, value) in enumerate(meta_rows, start=3):
        _set(ws_meta, i, 1, label, font=BLACK_BOLD, align=Alignment(vertical="top"))
        _set(ws_meta, i, 2, value, align=WRAP_TOP, border=BOX)
        ws_meta.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 40
    ws_meta.column_dimensions["C"].width = 20
    ws_meta.column_dimensions["D"].width = 20

    # ------ Aba 2: Bowtie visual ------
    ws = wb.create_sheet("Bowtie")

    # Larguras fixas das colunas
    widths = {
        "A": 35,  # Causa
        "B": 28,  # Barr Prev
        "C": 35,  # Ação Prev
        "D": 2,   # spacer
        "E": 30,  # Top Event
        "F": 2,   # spacer
        "G": 35,  # Impacto
        "H": 28,  # Barr Corr
        "I": 35,  # Ação Corr
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Título
    ws.merge_cells("A1:I1")
    _set(
        ws,
        1,
        1,
        f"Bowtie — {risco.codigo}: {risco.nome[:100]}",
        fill=HEADER_DARK,
        font=WHITE_BIG,
        align=WRAP_CENTER,
    )
    ws.row_dimensions[1].height = 28

    # Freq info
    freq_info = (
        f"Frequência PURA: {bowtie.frequencia_pura if bowtie else '-'}  "
        f"|  Frequência RESIDUAL: {bowtie.frequencia_residual if bowtie else '-'}  "
        f"|  Classificação residual: {risco.classificacao_residual or '-'}"
    )
    ws.merge_cells("A2:I2")
    _set(ws, 2, 1, freq_info, fill=HEADER_LIGHT, font=WHITE_BOLD, align=WRAP_CENTER)

    # Top event header centralizado (linha 4) — acima das colunas de dados
    _set(ws, 4, 5, "TOP EVENT", fill=TOP_EVENT_FILL, font=Font(bold=True, size=11), align=WRAP_CENTER, border=BOX)
    # Texto do top event na linha 5 (mergeado vertical até N-últimas linhas é difícil; fica estático nessa linha)
    top_event_text = (bowtie.top_event if bowtie else risco.nome)[:300]

    # Preparar linhas de causa com barreiras + ações
    # Cada causa ocupa max(1, nbarr) linhas
    data_start = 7  # header das colunas fica em 6

    # Headers das 5 colunas de dados
    _header_row(
        ws,
        6,
        [
            (1, "CAUSA (ameaça)"),
            (2, "BARREIRA PREVENTIVA"),
            (3, "AÇÃO PREVENTIVA"),
            (5, "RISCO / TOP EVENT"),
            (7, "IMPACTO (consequência)"),
            (8, "BARREIRA CORRETIVA"),
            (9, "AÇÃO CORRETIVA"),
        ],
    )

    # Distribuir ações preventivas aos nós das causas (uma por linha)
    causas = bowtie.causas if bowtie else []
    consequencias = bowtie.consequencias if bowtie else []

    # Construir lista de linhas para o lado esquerdo (causas):
    # Cada linha = (causa_idx_ou_None, barreira_texto, acao_texto, critica)
    left_rows: list[tuple[int | None, str, str, bool, str]] = []
    for ci, c in enumerate(causas):
        barrs = c.barreiras
        # Distribui ações preventivas deste risco entre as causas (round-robin por causa)
        acoes_para_causa = [
            a for i, a in enumerate(acoes_prev) if i % max(1, len(causas)) == ci
        ]
        n = max(1, len(barrs), len(acoes_para_causa))
        for j in range(n):
            barr_txt = barrs[j].descricao if j < len(barrs) else ""
            acao = acoes_para_causa[j] if j < len(acoes_para_causa) else None
            acao_txt = (
                f"{acao.codigo or ''}: {acao.descricao}"
                f"\nResp: {acao.responsavel.nome if acao.responsavel else '—'}"
                f" | Status: {acao.status}"
                + (f"\nPrazo: {acao.data_fim.isoformat()}" if acao and acao.data_fim else "")
                if acao
                else ""
            )
            left_rows.append(
                (ci if j == 0 else None, barr_txt, acao_txt, c.critica, c.codigo)
            )

    right_rows: list[tuple[int | None, str, str, bool, str]] = []
    for qi, q in enumerate(consequencias):
        barrs = q.barreiras
        acoes_para_cons = [
            a for i, a in enumerate(acoes_corr) if i % max(1, len(consequencias)) == qi
        ]
        n = max(1, len(barrs), len(acoes_para_cons))
        for j in range(n):
            barr_txt = barrs[j].descricao if j < len(barrs) else ""
            acao = acoes_para_cons[j] if j < len(acoes_para_cons) else None
            acao_txt = (
                f"{acao.codigo or ''}: {acao.descricao}"
                f"\nResp: {acao.responsavel.nome if acao.responsavel else '—'}"
                f" | Status: {acao.status}"
                if acao
                else ""
            )
            right_rows.append(
                (qi if j == 0 else None, barr_txt, acao_txt, q.critica, q.codigo)
            )

    total_rows = max(len(left_rows), len(right_rows), 1)

    # Top event centralizado mergeado em coluna E ocupando todas as linhas de dados
    top_start = data_start
    top_end = data_start + total_rows - 1
    ws.merge_cells(start_row=top_start, start_column=5, end_row=top_end, end_column=5)
    _set(
        ws,
        top_start,
        5,
        top_event_text,
        fill=TOP_EVENT_FILL,
        font=Font(bold=True, size=12),
        align=WRAP_CENTER,
        border=BOX,
    )

    # Preenche linhas esquerda e direita
    for i in range(total_rows):
        row_num = data_start + i
        # Esquerda
        if i < len(left_rows):
            causa_idx, barr, acao, crit, c_codigo = left_rows[i]
            if causa_idx is not None:
                c = causas[causa_idx]
                # Contar quantas linhas esta causa ocupa
                span = sum(1 for r in left_rows if r[0] == causa_idx) + sum(
                    1 for r in left_rows[i + 1 :] if r[0] is None and r[4] == c_codigo
                )
                # Mais simples: já que registramos causa_idx só na primeira linha do grupo,
                # contar linhas consecutivas com causa_idx None até achar outra com causa_idx ou fim.
                span = 1
                for k in range(i + 1, len(left_rows)):
                    if left_rows[k][0] is None:
                        span += 1
                    else:
                        break
                if span > 1:
                    ws.merge_cells(
                        start_row=row_num, start_column=1, end_row=row_num + span - 1, end_column=1
                    )
                border = CRITICAL_BORDER if crit else BOX
                _set(
                    ws,
                    row_num,
                    1,
                    f"[{c.codigo}]{' ⚠CRÍTICA' if crit else ''}\n{c.descricao}",
                    fill=CAUSA_FILL,
                    align=WRAP_TOP,
                    border=border,
                )
            _set(ws, row_num, 2, barr, fill=BARRIER_PREV_FILL, align=WRAP_TOP, border=BOX)
            _set(ws, row_num, 3, acao, fill=ACAO_PREV_FILL, align=WRAP_TOP, border=BOX)
        else:
            _set(ws, row_num, 2, "", border=BOX)
            _set(ws, row_num, 3, "", border=BOX)

        # Direita
        if i < len(right_rows):
            cons_idx, barr, acao, crit, q_codigo = right_rows[i]
            if cons_idx is not None:
                q = consequencias[cons_idx]
                span = 1
                for k in range(i + 1, len(right_rows)):
                    if right_rows[k][0] is None:
                        span += 1
                    else:
                        break
                if span > 1:
                    ws.merge_cells(
                        start_row=row_num, start_column=7, end_row=row_num + span - 1, end_column=7
                    )
                border = CRITICAL_BORDER if crit else BOX
                _set(
                    ws,
                    row_num,
                    7,
                    f"[{q.codigo}]{' ⚠CRÍTICA' if crit else ''}\n{q.descricao}",
                    fill=CONS_FILL,
                    align=WRAP_TOP,
                    border=border,
                )
            _set(ws, row_num, 8, barr, fill=BARRIER_CORR_FILL, align=WRAP_TOP, border=BOX)
            _set(ws, row_num, 9, acao, fill=ACAO_CORR_FILL, align=WRAP_TOP, border=BOX)
        else:
            _set(ws, row_num, 8, "", border=BOX)
            _set(ws, row_num, 9, "", border=BOX)

        ws.row_dimensions[row_num].height = 70

    # ------ Aba 3: Lista detalhada de Ações ------
    ws_ac = wb.create_sheet("Ações detalhadas")
    ac_headers = [
        "Cód.",
        "Descrição",
        "Tipo",
        "Status",
        "%",
        "Dono do Risco",
        "Responsável",
        "Área",
        "Categoria",
        "Grupo de trabalho",
        "Data início",
        "Data fim",
        "Detalhamento",
    ]
    for j, h in enumerate(ac_headers, start=1):
        _set(ws_ac, 1, j, h, fill=HEADER_DARK, font=WHITE_BOLD, align=WRAP_CENTER, border=BOX)
    ws_ac.row_dimensions[1].height = 22

    todas_acoes = acoes_prev + acoes_corr
    for i, a in enumerate(todas_acoes, start=2):
        _set(ws_ac, i, 1, a.codigo or "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 2, a.descricao, align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 3, a.tipo, align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 4, a.status, align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 5, a.percentual, align=WRAP_CENTER, border=BOX)
        _set(ws_ac, i, 6, a.dono_risco.nome if a.dono_risco else "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 7, a.responsavel.nome if a.responsavel else "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 8, a.area or "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 9, a.categoria or "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 10, a.grupo_trabalho or "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 11, a.data_inicio.isoformat() if a.data_inicio else "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 12, a.data_fim.isoformat() if a.data_fim else "", align=WRAP_TOP, border=BOX)
        _set(ws_ac, i, 13, a.detalhamento or "", align=WRAP_TOP, border=BOX)
        ws_ac.row_dimensions[i].height = 40

    for col, w in zip("ABCDEFGHIJKLM", [8, 50, 12, 16, 6, 22, 22, 20, 22, 28, 12, 12, 50]):
        ws_ac.column_dimensions[col].width = w

    # Ativa autofilter
    ws_ac.auto_filter.ref = f"A1:M{len(todas_acoes) + 1 if todas_acoes else 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
