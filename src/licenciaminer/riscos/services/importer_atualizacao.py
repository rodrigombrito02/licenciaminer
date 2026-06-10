"""Importer da planilha de atualização 2024 (Controles + Plano de Ações).

Arquivo: `Riscos Estratégicos Compactos_2024.09.07 - Atualização Riscoslimpa.xlsx`
Abas relevantes:
- Controles: 167 linhas (124 preventivos + 43 corretivos) — lista master atualizada
- Plano de Ações: 276 linhas (215 preventivas + 46 corretivas)

Estratégia: substitui os Controles e Acoes existentes dos riscos MUSA pela versão
detalhada desta planilha. Vincula pelo campo "ID MUSA" (ex: EST-CLO-FEL3-6).
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from licenciaminer.riscos.models import Acao, Controle, Pessoa, Risco

logger = logging.getLogger(__name__)

DEFAULT_ATUALIZACAO_PATH = (
    Path(__file__).resolve().parents[4].parent
    / "Riscos"
    / "Riscos Estratégicos Compactos_2024.09.07 - Atualização Riscoslimpa.xlsx"
)


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _get_or_create_pessoa(
    db: Session, nome: str, area: str | None = None
) -> Pessoa | None:
    nome = nome.strip()
    if not nome:
        return None
    p = db.query(Pessoa).filter_by(nome=nome).first()
    if p:
        return p
    p = Pessoa(nome=nome, area=area)
    db.add(p)
    db.flush()
    return p


def _status_to_canonical(texto: str) -> str:
    """Normaliza status do Excel para valores enumerados."""
    t = texto.lower()
    if "concl" in t:
        return "concluida"
    if "em andamento" in t:
        return "em_andamento"
    if "detalhamento" in t:
        return "detalhamento_em_andamento"
    if "a iniciar" in t or "não iniciada" in t or "nao iniciada" in t:
        return "nao_iniciada"
    if "cancelada" in t:
        return "cancelada"
    if "atrasada" in t:
        return "atrasada"
    # Valor preservado como "outro"
    return texto[:60] if texto else "nao_iniciada"


def _tipo_controle_norm(texto: str) -> str:
    t = texto.lower().strip()
    if t.startswith("prev"):
        return "preventivo"
    if t.startswith("corr"):
        return "corretivo"
    return "preventivo"


def _tipo_acao_norm(texto: str) -> str:
    t = texto.lower().strip()
    if t.startswith("prev"):
        return "preventiva"
    if t.startswith("corr"):
        return "corretiva"
    return "preventiva"


def _estagio_novo(texto: str) -> str | None:
    t = texto.lower().strip()
    if "implanta" in t or "transi" in t:
        return "implantacao"
    if "opera" in t:
        return "operacao"
    if "aprova" in t or "fel" in t:
        return "aprovacao"
    return None


def importar_atualizacao(
    db: Session, arquivo: Path | str = DEFAULT_ATUALIZACAO_PATH
) -> dict[str, int]:
    """Importa Controles e Plano de Ações detalhados, substituindo os existentes."""
    path = Path(arquivo)
    if not path.exists():
        logger.warning(f"Arquivo de atualização não encontrado: {path}")
        return {"controles": 0, "acoes": 0, "pessoas_criadas": 0}

    wb = load_workbook(path, data_only=True)

    # Mapa de riscos por ID MUSA
    riscos_por_cod: dict[str, Risco] = {r.codigo: r for r in db.query(Risco).all()}
    if not riscos_por_cod:
        logger.warning("Nenhum risco cadastrado; rode o importer MUSA primeiro.")
        return {"controles": 0, "acoes": 0, "pessoas_criadas": 0}

    # Apaga controles e ações existentes dos riscos MUSA para substituir
    db.query(Controle).delete()
    db.query(Acao).delete()
    db.commit()

    pessoas_antes = db.query(Pessoa).count()

    # ---- CONTROLES ----
    ws = wb["Controles"]
    ctrl_count = 0
    for row in ws.iter_rows(min_row=3, max_row=500, values_only=True):
        if not row or not any(row[:8]):
            continue
        id_musa = _norm(row[0])
        descricao = _norm(row[6])
        tipo_raw = _norm(row[7])
        if not id_musa or not descricao:
            continue
        risco = riscos_por_cod.get(id_musa)
        if not risco:
            continue
        categoria = _norm(row[5]) or None
        responsavel_nome = _norm(row[8])
        comentarios = _norm(row[9]) or None
        status_raw = _norm(row[10])
        resp = _get_or_create_pessoa(db, responsavel_nome, area=categoria)
        ctrl = Controle(
            risco_id=risco.id,
            bowtie_id=risco.bowties[0].id if risco.bowties else None,
            descricao=descricao,
            tipo=_tipo_controle_norm(tipo_raw),
            responsavel_id=resp.id if resp else None,
            categoria=categoria,
            comentarios=comentarios,
            status_teste=status_raw.lower()[:30] if status_raw else None,
        )
        db.add(ctrl)
        ctrl_count += 1

    # ---- PLANO DE AÇÕES ----
    ws = wb["Plano de Ações"]
    ac_count = 0
    # Cache de riscos NOVO criados (por descrição curta) para agrupar ações
    novos_riscos_cache: dict[str, Risco] = {}
    next_novo_idx = 1
    for row in ws.iter_rows(min_row=3, max_row=1050, values_only=True):
        if not row or not any(row[:15]):
            continue
        id_musa = _norm(row[1])
        acao_desc = _norm(row[10])
        if not acao_desc:
            continue
        risco: Risco | None = riscos_por_cod.get(id_musa)
        if not risco:
            # Caso especial: id_musa='NOVO' (risco novo fora do MUSA original).
            # Agrupa por descrição do risco (col 4); cria risco novo se não existir.
            if id_musa == "NOVO":
                risco_desc_novo = _norm(row[4])
                if not risco_desc_novo:
                    continue
                # Procura no cache de novos já criados
                risco = novos_riscos_cache.get(risco_desc_novo)
                if not risco:
                    # Procura no banco (caso execução anterior já tenha criado)
                    codigo_candidato = f"NOVO-{next_novo_idx:02d}"
                    while db.query(Risco).filter_by(codigo=codigo_candidato).first():
                        next_novo_idx += 1
                        codigo_candidato = f"NOVO-{next_novo_idx:02d}"
                    # Reusa risco existente por nome se possível
                    existente = (
                        db.query(Risco).filter_by(nome=risco_desc_novo[:300]).first()
                    )
                    if existente:
                        risco = existente
                    else:
                        risco = Risco(
                            codigo=codigo_candidato,
                            nome=risco_desc_novo[:300],
                            descricao=risco_desc_novo,
                            estagio=_estagio_novo(_norm(row[5])),
                        )
                        db.add(risco)
                        db.flush()
                        next_novo_idx += 1
                    novos_riscos_cache[risco_desc_novo] = risco
                    riscos_por_cod[risco.codigo] = risco
            else:
                # id_musa não encontrado e não é NOVO → ignora
                continue
        codigo_acao = _norm(row[9]) or _norm(row[0])
        etapa = _norm(row[5])
        categoria = _norm(row[6])
        dono_nome = _norm(row[7])
        subrisco = _norm(row[8]) or None
        status_raw = _norm(row[11])
        grupo_trabalho = _norm(row[12]) or None
        tema_relacionado = _norm(row[13]) or None
        tipo_tratamento = _norm(row[18])
        responsavel_nome = _norm(row[19])
        data_inicio = _to_date(row[20])
        data_fim = _to_date(row[21])
        inicio_txt = _norm(row[23]) or None
        conclusao_txt = _norm(row[24]) or None
        detalhamento = _norm(row[25]) or None
        valor = _to_float(row[26])
        evidencias = _norm(row[27]) or None

        dono = _get_or_create_pessoa(db, dono_nome, area=categoria)
        responsavel = _get_or_create_pessoa(db, responsavel_nome, area=categoria)

        percentual = 100 if "conclu" in status_raw.lower() else (50 if "em andamento" in status_raw.lower() else 0)
        status_norm = _status_to_canonical(status_raw)

        acao = Acao(
            risco_id=risco.id,
            bowtie_id=risco.bowties[0].id if risco.bowties else None,
            codigo=codigo_acao,
            descricao=acao_desc,
            tipo=_tipo_acao_norm(tipo_tratamento),
            responsavel_id=responsavel.id if responsavel else (dono.id if dono else None),
            dono_risco_id=dono.id if dono else None,
            area=responsavel_nome or categoria,
            categoria=categoria or etapa,
            subrisco=subrisco,
            grupo_trabalho=grupo_trabalho,
            tema_relacionado=tema_relacionado,
            prazo=data_fim,
            data_inicio=data_inicio,
            data_fim=data_fim,
            inicio_texto=inicio_txt,
            conclusao_texto=conclusao_txt,
            status=status_norm,
            percentual=percentual,
            detalhamento=detalhamento,
            valor_estimado=valor,
            evidencias=evidencias,
        )
        db.add(acao)
        ac_count += 1

    db.commit()
    pessoas_criadas = db.query(Pessoa).count() - pessoas_antes
    logger.info(
        f"Atualização import: +{ctrl_count} controles, +{ac_count} ações, "
        f"+{pessoas_criadas} pessoas"
    )
    return {
        "controles": ctrl_count,
        "acoes": ac_count,
        "pessoas_criadas": pessoas_criadas,
    }
