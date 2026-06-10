"""Importer do Excel MUSA (Bowtie Riscos Estratégicos).

Lê a aba "Riscos Estratégicos" (lista mestre) e as abas "Bow-Tie - R.XX" (bowties
detalhados), e popula o SQLite com Risco + Bowtie + Causa + Consequencia +
BarreiraPreventiva + BarreiraCorretiva.

Mapeamento MUSA 4×4 → Ternium 5×5:
- Impacto: B(Baixo)→1, M(Moderado)→3, A(Alto)→4, C(Crítico)→5
- Probabilidade: I(Improvável)→1, PP(Pouco Provável)→2, P(Provável)→4, F(Frequente)→5
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from licenciaminer.riscos.models import (
    Acao,
    BarreiraCorretiva,
    BarreiraPreventiva,
    Bowtie,
    Categoria,
    Causa,
    Consequencia,
    Controle,
    EloCadeiaValor,
    Pessoa,
    Risco,
    UnidadeOrg,
)
from licenciaminer.riscos.services.riscos import recalcular_classificacoes

logger = logging.getLogger(__name__)

DEFAULT_MUSA_PATH = (
    Path(__file__).resolve().parents[4].parent
    / "Riscos"
    / "Bowtie Riscos Estratégicos MUSA 20 11 apresentacoes.xlsx"
)

IMPACTO_MUSA_TO_TERNIUM = {
    "B": 1, "BAIXO": 1,
    "M": 3, "MODERADO": 3,
    "A": 4, "ALTO": 4,
    "C": 5, "CRITICO": 5, "CRÍTICO": 5,
}

PROB_MUSA_TO_TERNIUM = {
    "I": 1, "IMPROVAVEL": 1, "IMPROVÁVEL": 1,
    "PP": 2, "POUCO PROVAVEL": 2, "POUCO PROVÁVEL": 2,
    "P": 4, "PROVAVEL": 4, "PROVÁVEL": 4,
    "F": 5, "FREQUENTE": 5,
}

LISTA_ABA = "Riscos Estratégicos"
LISTA_HEADER_ROW = 15
LISTA_DATA_START = 16

# Colunas da aba "Riscos Estratégicos" (1-based)
COL_SEQ = 2
COL_AREA = 3
COL_COD_AREA = 4
COL_SUB_AREA = 5
COL_ETAPA = 6
COL_FASE = 7
COL_CODIGO = 8
COL_CATEGORIA = 9
COL_DESCRICAO = 10
COL_CAUSAS = 11
COL_CONSEQUENCIA = 12
# Primeira avaliação (residual): cols 13-17 (Impacto, Prob, C1, C2, Qualif)
COL_IMPACTO_R = 13
COL_PROB_R = 14
# Última avaliação (pura): cols 28-32
COL_IMPACTO_P = 28
COL_PROB_P = 29
# Responsáveis e mitigação
COL_RESPONSAVEL_ID = 33
COL_ESTRATEGIA = 47
COL_CONTROLES_EXISTENTES = 48
COL_ACAO_MITIGACAO = 49
COL_RESP_ACAO = 52
COL_DATA_IMPLANTACAO = 53
COL_SITUACAO = 55


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _map_impacto(v: Any) -> int | None:
    s = _norm(v).upper()
    if not s:
        return None
    if s.isdigit():
        n = int(s)
        return n if 1 <= n <= 5 else None
    return IMPACTO_MUSA_TO_TERNIUM.get(s)


def _map_prob(v: Any) -> int | None:
    s = _norm(v).upper()
    if not s:
        return None
    if s.isdigit():
        n = int(s)
        return n if 1 <= n <= 5 else None
    return PROB_MUSA_TO_TERNIUM.get(s)


def _get_or_create_categoria(db: Session, nome: str) -> Categoria | None:
    if not nome:
        return None
    cat = db.query(Categoria).filter_by(nome=nome).first()
    if cat:
        return cat
    cat = Categoria(nome=nome, descricao=f"Importado do MUSA: {nome}", cor="#64748b")
    db.add(cat)
    db.flush()
    return cat


def _get_or_create_pessoa(db: Session, nome: str, area: str | None = None) -> Pessoa | None:
    if not nome:
        return None
    p = db.query(Pessoa).filter_by(nome=nome).first()
    if p:
        return p
    p = Pessoa(nome=nome, area=area)
    db.add(p)
    db.flush()
    return p


def _infer_unidade_org(db: Session, area_nome: str) -> UnidadeOrg | None:
    """Match heurístico: procura UnidadeOrg cujo nome contenha a área do MUSA."""
    if not area_nome:
        return None
    termos = area_nome.lower().split()
    unidades = db.query(UnidadeOrg).all()
    for u in unidades:
        un = u.nome.lower()
        if any(t in un for t in termos if len(t) > 3):
            return u
    return None


def _infer_elo_cadeia(db: Session, categoria: str, area: str) -> EloCadeiaValor | None:
    texto = f"{categoria} {area}".lower()
    mapping = {
        "licenciamento": "Meio Ambiente e Licenciamento",
        "ambiental": "Meio Ambiente e Licenciamento",
        "comunidade": "Relações com Comunidades",
        "barragem": "Gestão de Barragens",
        "lavra": "Lavra",
        "beneficiamento": "Beneficiamento",
        "log": "Logística e Escoamento",
        "comercial": "Comercialização",
        "seguranca": "Saúde, Segurança e Emergências",
        "segurança": "Saúde, Segurança e Emergências",
        "pesquisa": "Pesquisa Mineral",
        "ti ": "TI e Dados",
        "financ": "Financeiro e Controladoria",
        "jur": "Jurídico e Compliance",
        "compliance": "Jurídico e Compliance",
        "pessoas": "Gestão de Pessoas",
    }
    for termo, nome_elo in mapping.items():
        if termo in texto:
            elo = db.query(EloCadeiaValor).filter_by(nome=nome_elo).first()
            if elo:
                return elo
    return None


def _codigo_to_sheetname(codigo: str) -> str | None:
    """Converte 'EST-REC-EXEC-1' → 'Bow-Tie - R.01' tentando casar pelo sufixo."""
    import re

    m = re.search(r"-(\d+)$", codigo)
    if not m:
        return None
    num = int(m.group(1))
    return f"Bow-Tie - R.{num:02d}"


def _parse_lista_estrategicos(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=LISTA_DATA_START, values_only=True):
        codigo = _norm(row[COL_CODIGO - 1])
        if not codigo:
            continue
        rows.append(
            {
                "codigo": codigo,
                "area": _norm(row[COL_AREA - 1]),
                "cod_area": _norm(row[COL_COD_AREA - 1]),
                "etapa": _norm(row[COL_ETAPA - 1]),
                "fase": _norm(row[COL_FASE - 1]),
                "categoria": _norm(row[COL_CATEGORIA - 1]),
                "descricao": _norm(row[COL_DESCRICAO - 1]),
                "causas": _norm(row[COL_CAUSAS - 1]),
                "consequencia": _norm(row[COL_CONSEQUENCIA - 1]),
                "impacto_residual": _map_impacto(row[COL_IMPACTO_R - 1]),
                "prob_residual": _map_prob(row[COL_PROB_R - 1]),
                "impacto_puro": _map_impacto(row[COL_IMPACTO_P - 1]),
                "prob_puro": _map_prob(row[COL_PROB_P - 1]),
                "estrategia": _norm(row[COL_ESTRATEGIA - 1]),
                "controles_existentes": _norm(row[COL_CONTROLES_EXISTENTES - 1]),
                "acao_mitigacao": _norm(row[COL_ACAO_MITIGACAO - 1]),
                "responsavel_acao": _norm(row[COL_RESP_ACAO - 1]),
                "situacao": _norm(row[COL_SITUACAO - 1]),
            }
        )
    return rows


def _estagio_from_etapa(etapa: str) -> str | None:
    e = etapa.upper()
    if e in ("EXEC", "EXECUTIVA", "EXECUÇÃO"):
        return "implantacao"
    if e in ("OPER", "OPERAÇÃO", "OPERACAO"):
        return "operacao"
    if e in ("FEL3", "COMS", "APROVACAO", "APROVAÇÃO"):
        return "aprovacao"
    return None


def _parse_bowtie_sheet(ws) -> dict[str, Any]:
    """Extrai causas, consequências, controles e ações da aba Bow-Tie."""
    causas: list[dict[str, str]] = []
    consequencias: list[dict[str, str]] = []
    controles_preventivos: list[dict[str, str]] = []
    controles_corretivos: list[dict[str, str]] = []
    acoes: list[dict[str, str]] = []

    # Metadados nas linhas 8-15 (col 2=label, col 3=valor)
    meta: dict[str, str] = {}
    for r in range(8, 16):
        label = _norm(ws.cell(row=r, column=2).value)
        val = _norm(ws.cell(row=r, column=3).value)
        if label and val:
            meta[label] = val

    contexto = _norm(ws.cell(row=8, column=9).value) or _norm(ws.cell(row=9, column=9).value)
    if contexto:
        meta["descricao_contexto"] = contexto

    # Causas + consequências (linhas 18-27)
    for r in range(18, 28):
        c_cod = _norm(ws.cell(row=r, column=4).value)
        c_desc = _norm(ws.cell(row=r, column=5).value)
        if c_cod and c_desc:
            causas.append({"codigo": c_cod, "descricao": c_desc})
        i_cod = _norm(ws.cell(row=r, column=19).value)
        i_desc = _norm(ws.cell(row=r, column=20).value)
        if i_cod and i_desc:
            consequencias.append({"codigo": i_cod, "descricao": i_desc})

    # Controles: linhas 30+ até encontrar "AÇÕES DE MELHORIA" ou ficar tudo vazio.
    # Depois disso vêm as ações.
    in_actions = False
    for r in range(30, ws.max_row + 1):
        label4 = _norm(ws.cell(row=r, column=4).value)
        if label4.upper().startswith("AÇÕES DE MELHORIA") or label4.upper().startswith(
            "ACOES DE MELHORIA"
        ):
            in_actions = True
            continue
        if not in_actions:
            cp_cod = _norm(ws.cell(row=r, column=4).value)
            cp_desc = _norm(ws.cell(row=r, column=5).value)
            if cp_cod.startswith("CP.") and cp_desc:
                controles_preventivos.append(
                    {
                        "codigo": cp_cod,
                        "descricao": cp_desc,
                        "eficacia": _norm(ws.cell(row=r, column=8).value),
                        "causas_ref": _norm(ws.cell(row=r, column=10).value),
                        "responsavel": _norm(ws.cell(row=r, column=11).value),
                        "status": _norm(ws.cell(row=r, column=2).value),
                        "comentario": _norm(ws.cell(row=r, column=3).value),
                    }
                )
            cc_cod = _norm(ws.cell(row=r, column=14).value)
            cc_desc = _norm(ws.cell(row=r, column=15).value)
            if cc_cod.startswith("CC.") and cc_desc:
                controles_corretivos.append(
                    {
                        "codigo": cc_cod,
                        "descricao": cc_desc,
                        "eficacia": _norm(ws.cell(row=r, column=18).value),
                        "impactos_ref": _norm(ws.cell(row=r, column=21).value),
                        "responsavel": _norm(ws.cell(row=r, column=22).value),
                        "status": _norm(ws.cell(row=r, column=23).value),
                        "comentario": _norm(ws.cell(row=r, column=24).value),
                    }
                )
        else:
            num = _norm(ws.cell(row=r, column=4).value)
            desc = _norm(ws.cell(row=r, column=5).value)
            controle_assoc = _norm(ws.cell(row=r, column=8).value)
            ref = _norm(ws.cell(row=r, column=9).value)
            prazo = _norm(ws.cell(row=r, column=10).value)
            responsavel = _norm(ws.cell(row=r, column=11).value)
            comentario = _norm(ws.cell(row=r, column=13).value)
            if num and desc:
                acoes.append(
                    {
                        "numero": num,
                        "descricao": desc,
                        "controle_assoc": controle_assoc,
                        "referencias": ref,
                        "prazo": prazo,
                        "responsavel": responsavel,
                        "comentario": comentario,
                    }
                )

    return {
        "meta": meta,
        "causas": causas,
        "consequencias": consequencias,
        "controles_preventivos": controles_preventivos,
        "controles_corretivos": controles_corretivos,
        "acoes": acoes,
    }


def _eficacia_int(texto: str) -> int | None:
    if not texto:
        return None
    s = texto.strip()
    if s.isdigit():
        n = int(s)
        return n if 1 <= n <= 5 else None
    return None


def importar_musa(db: Session, arquivo: Path | str = DEFAULT_MUSA_PATH) -> dict[str, int]:
    """Carrega MUSA e popula riscos + bowties. Idempotente (skip por código)."""
    path = Path(arquivo)
    if not path.exists():
        logger.warning(f"MUSA não encontrado: {path}")
        return {"riscos": 0, "bowties": 0, "causas": 0, "consequencias": 0, "barreiras": 0}

    wb = load_workbook(path, data_only=True)
    ws_lista = wb[LISTA_ABA]
    linhas = _parse_lista_estrategicos(ws_lista)

    r_count = 0
    b_count = 0
    cs_count = 0
    cn_count = 0
    br_count = 0
    ctrl_count = 0
    ac_count = 0

    for linha in linhas:
        codigo = linha["codigo"]
        if db.query(Risco).filter_by(codigo=codigo).first():
            continue

        categoria = _get_or_create_categoria(db, linha["categoria"])
        responsavel = _get_or_create_pessoa(
            db, linha["responsavel_acao"].split("\n")[0].replace("GG responsável:", "").strip()
            if linha["responsavel_acao"]
            else "",
            area=linha["area"],
        )
        unidade = _infer_unidade_org(db, linha["area"])
        elo = _infer_elo_cadeia(db, linha["categoria"], linha["area"])

        metadados = {
            "cod_area": linha["cod_area"],
            "etapa_musa": linha["etapa"],
            "fase": linha["fase"],
            "estrategia_mitigacao": linha["estrategia"],
            "controles_existentes": linha["controles_existentes"],
            "acao_mitigacao": linha["acao_mitigacao"],
            "situacao": linha["situacao"],
        }

        risco = Risco(
            codigo=codigo,
            nome=linha["descricao"][:300] if linha["descricao"] else codigo,
            descricao=linha["descricao"],
            estagio=_estagio_from_etapa(linha["etapa"]),
            categoria_id=categoria.id if categoria else None,
            responsavel_id=responsavel.id if responsavel else None,
            unidade_org_id=unidade.id if unidade else None,
            elo_cadeia_valor_id=elo.id if elo else None,
            prob_pura=linha["prob_puro"],
            impacto_pura=linha["impacto_puro"],
            prob_residual=linha["prob_residual"],
            impacto_residual=linha["impacto_residual"],
            metadados_json=json.dumps(metadados, ensure_ascii=False),
        )
        recalcular_classificacoes(db, risco)
        db.add(risco)
        db.flush()
        r_count += 1

        # Bowtie a partir da aba correspondente, se existir
        sheet_name = _codigo_to_sheetname(codigo)
        if sheet_name and sheet_name in wb.sheetnames:
            parsed = _parse_bowtie_sheet(wb[sheet_name])
            bowtie = Bowtie(
                risco_id=risco.id,
                top_event=parsed["meta"].get("Risco") or linha["descricao"][:300],
                hazard=parsed["meta"].get("descricao_contexto"),
                frequencia_pura=linha["prob_puro"],
                frequencia_residual=linha["prob_residual"],
            )
            db.add(bowtie)
            db.flush()
            b_count += 1

            # Cria Controles (entidade) primeiro para poder vincular BarreiraPreventiva.controle_id
            cp_objs_by_codigo: dict[str, Controle] = {}
            for cp in parsed["controles_preventivos"]:
                resp = _get_or_create_pessoa(db, cp["responsavel"])
                ctrl = Controle(
                    risco_id=risco.id,
                    bowtie_id=bowtie.id,
                    descricao=f"{cp['codigo']}: {cp['descricao']}",
                    tipo="preventivo",
                    responsavel_id=resp.id if resp else None,
                    efetividade=_eficacia_int(cp["eficacia"]),
                    status_teste=(cp["status"].lower() if cp["status"] else None),
                )
                db.add(ctrl)
                db.flush()
                cp_objs_by_codigo[cp["codigo"]] = ctrl
                ctrl_count += 1

            cc_objs_by_codigo: dict[str, Controle] = {}
            for cc in parsed["controles_corretivos"]:
                resp = _get_or_create_pessoa(db, cc["responsavel"])
                ctrl = Controle(
                    risco_id=risco.id,
                    bowtie_id=bowtie.id,
                    descricao=f"{cc['codigo']}: {cc['descricao']}",
                    tipo="corretivo",
                    responsavel_id=resp.id if resp else None,
                    efetividade=_eficacia_int(cc["eficacia"]),
                    status_teste=(cc["status"].lower() if cc["status"] else None),
                )
                db.add(ctrl)
                db.flush()
                cc_objs_by_codigo[cc["codigo"]] = ctrl
                ctrl_count += 1

            # Causas — primeira marcada como crítica (para teste de regra)
            for ordem, c in enumerate(parsed["causas"]):
                causa = Causa(
                    bowtie_id=bowtie.id,
                    codigo=c["codigo"],
                    descricao=c["descricao"],
                    ordem=ordem,
                    critica=(ordem == 0),
                )
                db.add(causa)
                db.flush()
                cs_count += 1

                cod_num = c["codigo"].replace("C.", "").strip()
                for cp in parsed["controles_preventivos"]:
                    referencias = [x.strip() for x in cp["causas_ref"].split(",") if x.strip()]
                    if (
                        cod_num in referencias
                        or "todas" in cp["causas_ref"].lower()
                    ):
                        db.add(
                            BarreiraPreventiva(
                                causa_id=causa.id,
                                descricao=f"{cp['codigo']}: {cp['descricao']}",
                                efetividade=_eficacia_int(cp["eficacia"]),
                                controle_id=cp_objs_by_codigo[cp["codigo"]].id,
                            )
                        )
                        br_count += 1

            # Consequências — primeira marcada como crítica
            for ordem, q in enumerate(parsed["consequencias"]):
                conseq = Consequencia(
                    bowtie_id=bowtie.id,
                    codigo=q["codigo"],
                    descricao=q["descricao"],
                    ordem=ordem,
                    critica=(ordem == 0),
                )
                db.add(conseq)
                db.flush()
                cn_count += 1
                cod_num = q["codigo"].replace("I.", "").strip()
                # Tenta ligar CC cujo "impactos_ref" cite este índice
                linked = False
                for cc in parsed["controles_corretivos"]:
                    referencias = [
                        x.strip() for x in cc["impactos_ref"].split(",") if x.strip()
                    ]
                    if (
                        cod_num in referencias
                        or "todas" in cc["impactos_ref"].lower()
                    ):
                        db.add(
                            BarreiraCorretiva(
                                consequencia_id=conseq.id,
                                descricao=f"{cc['codigo']}: {cc['descricao']}",
                                efetividade=_eficacia_int(cc["eficacia"]),
                                controle_id=cc_objs_by_codigo[cc["codigo"]].id,
                            )
                        )
                        br_count += 1
                        linked = True
                # Se nenhum corretivo foi referenciado, distribui round-robin
                if not linked and parsed["controles_corretivos"]:
                    cc = parsed["controles_corretivos"][
                        ordem % len(parsed["controles_corretivos"])
                    ]
                    db.add(
                        BarreiraCorretiva(
                            consequencia_id=conseq.id,
                            descricao=f"{cc['codigo']}: {cc['descricao']}",
                            efetividade=_eficacia_int(cc["eficacia"]),
                            controle_id=cc_objs_by_codigo[cc["codigo"]].id,
                        )
                    )
                    br_count += 1

            # Ações de melhoria → Acao
            for a in parsed["acoes"]:
                resp = _get_or_create_pessoa(db, a["responsavel"])
                controle_assoc = a["controle_assoc"]
                tipo = "corretiva" if controle_assoc.startswith("CC.") else "preventiva"
                bowtie_id_ref = bowtie.id
                acao = Acao(
                    risco_id=risco.id,
                    bowtie_id=bowtie_id_ref,
                    descricao=f"#{a['numero']}: {a['descricao']}",
                    tipo=tipo,
                    responsavel_id=resp.id if resp else None,
                    comentario=(
                        f"Controle associado: {controle_assoc}"
                        + (f" | Refs: {a['referencias']}" if a["referencias"] else "")
                        + (f" | Prazo (texto): {a['prazo']}" if a["prazo"] else "")
                        + (f" | Obs: {a['comentario']}" if a["comentario"] else "")
                    ).strip(),
                )
                db.add(acao)
                ac_count += 1

    db.commit()
    logger.info(
        f"MUSA import: +{r_count} riscos, +{b_count} bowties, "
        f"+{cs_count} causas, +{cn_count} consequências, +{br_count} barreiras, "
        f"+{ctrl_count} controles, +{ac_count} ações"
    )
    return {
        "riscos": r_count,
        "bowties": b_count,
        "causas": cs_count,
        "consequencias": cn_count,
        "barreiras": br_count,
        "controles": ctrl_count,
        "acoes": ac_count,
    }
